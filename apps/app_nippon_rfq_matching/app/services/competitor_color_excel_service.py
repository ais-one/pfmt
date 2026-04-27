"""
Service for parsing competitor color comparison from multisheet Excel files
"""

import logging
import re
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from apps.app_nippon_rfq_matching.app.models.competitor import CompetitorColorComparison

logger = logging.getLogger(__name__)


class CompetitorColorExcelService:
    """Service for parsing and storing competitor color comparison data from Excel files"""

    def parse_excel_to_cell_mapping(self, file_path: str) -> dict[str, dict[str, str]]:
        """
        Parse Excel file to create cell mapping for each sheet

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with sheet names as keys and cell mappings as values
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            result = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_data = {}

                for row in ws.iter_rows():
                    for cell in row:
                        # Skip jika None / kosong
                        if cell.value is None:
                            continue

                        # Ambil alamat cell (contoh: A1, B2)
                        cell_address = cell.coordinate

                        # Simpan value
                        sheet_data[cell_address] = str(cell.value)

                # Simpan per sheet
                result[sheet_name] = sheet_data

            logger.info(f"Processed {len(result)} sheets from Excel file: {file_path}")
            return result

        except Exception as e:
            logger.error(f"Error parsing Excel to cell mapping: {e}", exc_info=True)
            raise

    def extract_brand(self, sheet_name: str, cells: dict[str, str]) -> str:
        """
        Extract brand from sheet name or cell A1

        Args:
            sheet_name: Name of the sheet
            cells: Cell mapping for the sheet

        Returns:
            Brand name in uppercase
        """
        # 1. Dari sheet name
        brand = sheet_name.split("_vs_")[0]

        if brand:
            return brand.upper()

        # 2. Fallback dari A1
        title = cells.get("A1", "")
        match = re.match(r"(\w+)", title)
        if match:
            return match.group(1).upper()

        # 3. Try other common sheet name patterns
        if "vs" in sheet_name.lower():
            parts = sheet_name.lower().split("vs")
            if parts:
                return parts[0].strip().upper()

        # 4. Fallback to known brands in cells
        cells_text = " ".join(cells.values()).upper()
        for brand in ["JOTUN", "INTERNATIONAL PAINT", "HEMPEL", "RAL", "SIGMA", "PPG"]:
            if brand in cells_text:
                return brand

        logger.warning(f"Could not determine brand from sheet: {sheet_name}")
        return "UNKNOWN"

    def transform_to_rows_dynamic(
        self, cell_mapping: dict[str, dict[str, str]]
    ) -> list[dict[str, Any]]:
        """
        Transform cell mapping to rows format

        Args:
            cell_mapping: Cell mapping from parse_excel_to_cell_mapping

        Returns:
            List of color comparison dictionaries
        """
        results = []

        for sheet_name, cells in cell_mapping.items():
            brand = self.extract_brand(sheet_name, cells)

            logger.info(f"Processing sheet '{sheet_name}' for brand '{brand}'")

            row = 3  # Start from row 3 based on the provided code
            while True:
                a = cells.get(f"A{row}")
                b = cells.get(f"B{row}")
                c = cells.get(f"C{row}")

                # stop kalau kosong semua
                if not a and not b and not c:
                    break

                # skip row invalid
                if not a or not b:
                    row += 1
                    continue

                # Extract item number
                item_no = None
                if a:
                    digits = re.findall(r"\d+", str(a))
                    if digits:
                        try:
                            item_no = int(digits[0])
                        except ValueError:
                            continue

                if item_no is None:
                    row += 1
                    continue

                results.append(
                    {
                        "item_no": item_no,
                        "source_brand": brand,
                        "source_code": str(b).strip().upper(),
                        "npms_code": str(c).strip().upper() if c else None,
                        "raw_text": f"{b} -> {c}"
                        if b and c
                        else f"{b}"
                        if b
                        else f"{c}"
                        if c
                        else "",
                    }
                )

                row += 1

        logger.info(f"Total rows extracted: {len(results)}")
        return results

    def parse_excel_file(self, file_path: str) -> dict[str, list[dict[str, Any]]]:
        """
        Parse Excel file and return grouped by brand

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with brand as keys and list of items as values
        """
        try:
            # Step 1: Parse Excel to cell mapping
            cell_mapping = self.parse_excel_to_cell_mapping(file_path)

            # Step 2: Transform to rows
            all_rows = self.transform_to_rows_dynamic(cell_mapping)

            # Step 3: Group by brand
            result = {}
            for row in all_rows:
                brand = row["source_brand"]
                if brand not in result:
                    result[brand] = []
                result[brand].append(row)

            logger.info(f"Excel parsing completed. Brands found: {list(result.keys())}")
            return result

        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}", exc_info=True)
            raise

    def save_to_database(
        self,
        parsed_data: dict[str, list[dict[str, Any]]],
        uploaded_file_id: int,
        db: Session,
    ) -> dict[str, Any]:
        """
        Save parsed competitor color comparison data to database with duplicate handling

        Args:
            parsed_data: Parsed data from Excel
            uploaded_file_id: ID of the uploaded file
            db: Database session

        Returns:
            Dictionary with keys:
            - created: List of newly created CompetitorColorComparison records
            - duplicates: List of duplicate record info (item_no, source_brand, source_code)
            - skipped_count: Number of skipped items (empty/invalid)
        """
        from sqlalchemy.exc import IntegrityError

        db_records = []
        duplicates = []
        skipped_count = 0

        try:
            for brand, items in parsed_data.items():
                for item_data in items:
                    item_no = item_data.get("item_no")
                    source_brand = item_data.get("source_brand")
                    source_code = item_data.get("source_code")
                    npms_code = item_data.get("npms_code")

                    # Skip if missing required fields
                    if not all([item_no is not None, source_brand, source_code]):
                        logger.warning(
                            f"Skipping record with missing required fields: item_no={item_no}, "
                            f"source_brand={source_brand}, source_code={source_code}"
                        )
                        skipped_count += 1
                        continue

                    # Check if record already exists
                    existing = (
                        db.query(CompetitorColorComparison)
                        .filter(
                            CompetitorColorComparison.item_no == item_no,
                            CompetitorColorComparison.source_brand == source_brand,
                            CompetitorColorComparison.source_code == source_code,
                        )
                        .first()
                    )

                    if existing:
                        # Track duplicate
                        duplicates.append(
                            {
                                "item_no": item_no,
                                "source_brand": source_brand,
                                "source_code": source_code,
                                "existing_id": existing.id,
                            }
                        )
                        logger.debug(
                            f"Duplicate record found: item_no={item_no}, source_brand={source_brand}, "
                            f"source_code={source_code}"
                        )
                    else:
                        # Create new record
                        db_record = CompetitorColorComparison(
                            item_no=item_no,
                            source_brand=source_brand,
                            source_code=source_code,
                            npms_code=npms_code,
                            raw_text=item_data.get("raw_text"),
                            uploaded_file_id=uploaded_file_id,
                        )
                        db.add(db_record)
                        db_records.append(db_record)
                        logger.debug(
                            f"Created new record: item_no={item_no}, source_brand={source_brand}, "
                            f"source_code={source_code}"
                        )

            try:
                db.commit()

                # Refresh to get IDs
                for record in db_records:
                    db.refresh(record)

                logger.info(
                    f"Saved CompetitorColorComparison from Excel: created={len(db_records)}, "
                    f"duplicates={len(duplicates)}, skipped={skipped_count}"
                )

            except IntegrityError as e:
                db.rollback()
                logger.error(f"Integrity error saving CompetitorColorComparison: {e}")
                raise

            return {
                "created": db_records,
                "duplicates": duplicates,
                "skipped_count": skipped_count,
            }

        except IntegrityError as e:
            db.rollback()
            logger.error(f"Integrity error saving CompetitorColorComparison: {e}")
            raise
        except Exception as e:
            db.rollback()
            logger.error(f"Error saving to database: {e}", exc_info=True)
            raise


# Singleton instance
competitor_color_excel_service = CompetitorColorExcelService()
