"""
RFQ Indonesia Excel Parser

Utility functions for parsing Indonesian RFQ Excel files (Sinar Bukittinggi, Solo, Bajo, Bandung).
This parser handles Excel files with RFQ data in specific formats.
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Header keywords that identify Indonesian RFQ format
INDONESIA_HEADER_KEYWORDS = ["Part Number", "No", "Description", "Delivery time"]


def is_indonesia_rfqs(cell_map: dict, sheet_name: str | None = None) -> bool:
    """Check if the cell data represents an Indonesian RFQ format Excel file.

    Indonesian RFQ format characteristics:
    - Header section with "RFQ NO", "VENDOR NAME", "VENDOR Code", "PAYMENT TERMS", "CURRENCY"
    - Equipment section with "EQUIPMENT", "MAKER", "MODEL"
    - Table headers: "No", "Part Number", "Description", "QTY", "Unit", "Price", "Total", "Delivery time"
    - Used for Indonesian vessels (Sinar Bukittinggi, Solo, Bajo, Bandung)

    Args:
        cell_map: Dictionary mapping cell coordinates to values.
        sheet_name: Name of the sheet (optional, not used for identification).

    Returns:
        True if the cell data matches Indonesian RFQ format characteristics.
    """
    # Group cells by row
    rows = {}
    for cell, val in cell_map.items():
        if len(cell) >= 2:
            col = cell[0]
            row_num = cell[1:]
            if row_num.isdigit():
                row_idx = int(row_num)
                if row_idx not in rows:
                    rows[row_idx] = {}
                rows[row_idx][col] = val

    if not rows:
        return False

    # Vendor headers typically in column B
    vendor_headers = {
        "RFQ NO",
        "VENDOR NAME",
        "VENDOR Code",
        "PAYMENT TERMS",
        "CURRENCY",
        "AUTHENTICITY",
    }
    found_vendor_headers = set()

    # Equipment headers typically in column E
    equipment_headers = {"EQUIPMENT", "MAKER", "MODEL", "SERIAL NO."}
    found_equipment_headers = set()

    # Table headers (all in same row)
    table_headers = {
        "No",
        "Part Number",
        "Description",
        "QTY",
        "Unit",
        "Price",
        "Total",
        "Delivery time",
    }

    for row_idx, row in rows.items():
        # Check for vendor headers in column B
        b_val = str(row.get("B", "")).upper().strip()
        if b_val in vendor_headers:
            found_vendor_headers.add(b_val)

        # Check for equipment headers in column E
        e_val = str(row.get("E", "")).upper().strip()
        if e_val in equipment_headers:
            found_equipment_headers.add(e_val)

        # Check for complete table header row
        row_values = [str(v).upper().strip() if v else "" for v in row.values()]
        set(row_values)

        # Count matching table headers in this row
        table_header_count = sum(
            1 for h in table_headers if any(h in val for val in row_values)
        )

        # If row contains multiple table headers (No + Part Number + Description + Delivery time)
        if table_header_count >= 4:
            return True

    # Alternative detection: vendor headers + equipment headers
    if len(found_vendor_headers) >= 2 and len(found_equipment_headers) >= 1:
        return True

    return False


def create_cell_map(file_path: str) -> dict:
    """
    Create a mapping of all cell coordinates to their values.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary mapping cell coordinates (e.g., 'A1') to their values

    Example:
        >>> cell_map = create_cell_map("rfq.xlsx")
        >>> print(cell_map.get('A1'))
    """
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        cell_map = {}

        for row in ws.iter_rows():
            for cell in row:
                cell_map[cell.coordinate] = cell.value

        logger.debug(f"Created cell map with {len(cell_map)} cells from {file_path}")

        return cell_map

    except Exception as e:
        logger.error(f"Error creating cell map from {file_path}: {e}", exc_info=True)
        raise


def find_header_row(
    cell_map: dict, sheet, header_keywords: list[str] | None = None
) -> int:
    """
    Find the row containing the main table headers.

    Args:
        cell_map: Dictionary mapping cell coordinates to values
        sheet: openpyxl worksheet object
        header_keywords: List of keywords to identify the header row.
                        Defaults to ["Part Number", "No", "Description"]

    Returns:
        Row index (1-based) containing the headers, or None if not found

    Example:
        >>> header_row = find_header_row(cell_map, ws)
        >>> print(f"Headers found at row: {header_row}")
    """
    if header_keywords is None:
        header_keywords = ["Part Number", "No", "Description"]

    # Search for the row that contains header keywords
    for row_idx in range(1, sheet.max_row + 1):
        row_values = [
            cell_map.get(f"{col}{row_idx}") for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        ]
        row_values = [str(v) if v is not None else "" for v in row_values]

        # Check if this row contains multiple header keywords
        keyword_count = sum(
            1
            for keyword in header_keywords
            if any(keyword.lower() in str(val).lower() for val in row_values)
        )

        if keyword_count >= 2:
            logger.debug(f"Found header row at index: {row_idx}")
            return row_idx

    logger.warning("Could not find header row, defaulting to row 9")
    return 9


def parse_rfq_excel(
    file_path: str,
    header_row_index: int | None = None,
    drop_na_columns: bool = True,
    drop_na_rows: bool = True,
    required_column: str | None = None,
) -> pd.DataFrame:
    """
    Parse Indonesian RFQ Excel file and extract table data.

    This function:
    1. Loads the Excel file using openpyxl
    2. Extracts headers from the specified row (or auto-detects)
    3. Extracts data rows from the table
    4. Creates a pandas DataFrame
    5. Cleans the data by dropping empty columns/rows

    Args:
        file_path: Path to the Excel file
        header_row_index: Row index (1-based) containing headers.
                         If None, attempts auto-detection. Defaults to None.
        drop_na_columns: Whether to drop columns that are entirely empty
        drop_na_rows: Whether to drop rows that are entirely empty
        required_column: Column name that must not be null. Rows with null
                        values in this column will be dropped.
                        Defaults to "Delivery time (in working days)"

    Returns:
        Cleaned DataFrame with RFQ data

    Example:
        >>> df = parse_rfq_excel("rfq_sinar_bukittinggi.xlsx")
        >>> print(df.columns)
        >>> print(df.head())
    """
    try:
        # Load workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # Create cell map for header detection if needed
        if header_row_index is None:
            cell_map = create_cell_map(file_path)
            header_row_index = find_header_row(cell_map, ws)

        # Extract headers
        headers = [cell.value for cell in ws[header_row_index]]

        # Extract data rows
        data_rows = []
        for row_index in range(header_row_index + 1, ws.max_row + 1):
            row_values = [cell.value for cell in ws[row_index]]
            data_rows.append(row_values)

        # Create DataFrame
        df = pd.DataFrame(data_rows, columns=headers)

        # Clean up the DataFrame
        if drop_na_columns:
            df = df.dropna(axis=1, how="all")
            logger.debug(f"Dropped empty columns, remaining: {len(df.columns)}")

        if drop_na_rows:
            before_count = len(df)
            df = df.dropna(axis=0, how="all")
            logger.debug(f"Dropped {before_count - len(df)} empty rows")

        # Filter by required column if specified
        if required_column and required_column in df.columns:
            before_count = len(df)
            df = df.dropna(subset=[required_column])
            logger.debug(
                f"Dropped {before_count - len(df)} rows with missing {required_column}"
            )
        elif required_column:
            logger.warning(
                f"Required column '{required_column}' not found in DataFrame"
            )

        df = df.reset_index(drop=True)

        logger.info(
            f"Parsed RFQ Excel file: {len(df)} rows, {len(df.columns)} columns from {file_path}"
        )

        return df

    except Exception as e:
        logger.error(f"Error parsing RFQ Excel file {file_path}: {e}", exc_info=True)
        raise


def parse_rfq_excel_to_dict(
    file_path: str,
    header_row_index: int | None = None,
    required_column: str | None = None,
) -> list[dict]:
    """
    Parse Indonesian RFQ Excel file and return as list of dictionaries.

    This is a convenience function that converts the DataFrame result
    to a list of dicts for easier API integration.

    Args:
        file_path: Path to the Excel file
        header_row_index: Row index (1-based) containing headers
        required_column: Column name that must not be null

    Returns:
        List of dictionaries, one per row in the parsed table

    Example:
        >>> items = parse_rfq_excel_to_dict("rfq_sinar_bukittinggi.xlsx")
        >>> for item in items:
        ...     print(item.get("Part Number"), item.get("Description"))
    """
    df = parse_rfq_excel(
        file_path=file_path,
        header_row_index=header_row_index,
        required_column=required_column,
    )

    # Convert NaN to None for JSON serialization
    result = df.where(pd.notna(df), None).to_dict("records")

    return result


def batch_parse_rfq_excels(
    file_paths: list[str],
    header_row_index: int | None = None,
    required_column: str | None = None,
) -> pd.DataFrame:
    """
    Parse multiple Indonesian RFQ Excel files and combine into a single DataFrame.

    Args:
        file_paths: List of paths to Excel files
        header_row_index: Row index (1-based) containing headers
        required_column: Column name that must not be null

    Returns:
        Combined DataFrame with data from all files

    Example:
        >>> files = ["rfq1.xlsx", "rfq2.xlsx", "rfq3.xlsx"]
        >>> df = batch_parse_rfq_excels(files)
        >>> print(f"Total rows: {len(df)}")
    """
    dataframes = []

    for file_path in file_paths:
        try:
            df = parse_rfq_excel(
                file_path=file_path,
                header_row_index=header_row_index,
                required_column=required_column,
            )
            # Add source file column for tracking
            df["source_file"] = Path(file_path).name
            dataframes.append(df)

        except Exception as e:
            logger.error(f"Failed to parse {file_path}: {e}", exc_info=True)
            continue

    if not dataframes:
        logger.warning("No files were successfully parsed")
        return pd.DataFrame()

    # Combine all dataframes
    combined_df = pd.concat(dataframes, ignore_index=True)
    logger.info(f"Combined {len(dataframes)} files into {len(combined_df)} total rows")

    return combined_df
