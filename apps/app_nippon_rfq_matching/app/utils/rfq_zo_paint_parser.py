"""
RFQ ZO Paint Excel Parser

Utility functions for parsing ZO Paint RFQ Excel files.
This parser handles "store req. form" sheets with section-based grouping
and extracts paint-related information including colors, volumes, and RAL codes.
"""

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Color detection constants
COLOR_LIST = {
    "RED",
    "GREEN",
    "BLUE",
    "GREY",
    "GRAY",
    "YELLOW",
    "BLACK",
    "WHITE",
    "ORANGE",
    "ALUMINIUM",
}

IGNORE_TOKENS = {"A", "STD", "BASE"}

# Special product type keywords in Chinese
CHINESE_SPECIAL_KEYWORDS = ["稀释剂", "固化剂"]  # thinner, curing agent

# Expected sheet name for ZO Paint format
ZO_PAINT_SHEET_NAME = "store req. form"


def is_zo_paint_excel(cell_map: dict, sheet_name: str | None = None) -> bool:
    """Check if the cell data represents a ZO Paint format Excel file.

    ZO Paint format characteristics:
    - Sheet name contains "store req. form" or similar
    - Has numeric values in column A (item numbers)
    - Has text values in column C (descriptions)
    - May contain paint-related keywords or Chinese characters

    Args:
        cell_map: Dictionary mapping cell coordinates to values.
        sheet_name: Name of the sheet (optional, helps with identification).

    Returns:
        True if the cell data matches ZO Paint format characteristics.
    """
    # Check sheet name if provided
    if sheet_name:
        sheet_lower = sheet_name.lower()
        if "store" in sheet_lower and "req" in sheet_lower:
            return True

    # Group cells by row
    rows = defaultdict(dict)
    for cell, val in cell_map.items():
        if len(cell) >= 2:
            col = cell[0]
            row_num = cell[1:]
            if row_num.isdigit():
                rows[int(row_num)][col] = val

    if not rows:
        return False

    # Check for ZO Paint structure:
    # - Column A contains numeric values (item numbers)
    # - Column C contains text (descriptions)
    # - At least 3 rows with this structure
    numeric_a_count = 0
    text_c_count = 0

    for row in rows.values():
        a_val = row.get("A")
        c_val = row.get("C")

        if isinstance(a_val, int | float) and a_val > 0:
            numeric_a_count += 1

        if c_val and isinstance(c_val, str) and len(c_val.strip()) > 0:
            text_c_count += 1

    # ZO Paint format should have multiple rows with numeric A and text C values
    return numeric_a_count >= 3 and text_c_count >= 3


def create_cell_map(
    file_path: str, sheet_name: str | None = None, include_none: bool = False
) -> dict:
    """
    Create a mapping of cell coordinates to their values.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to load. If None, uses the first sheet
        include_none: Whether to include None values in the map. Defaults to False

    Returns:
        Dictionary mapping cell coordinates (e.g., 'A1') to their values

    Example:
        >>> cell_map = create_cell_map("zo_paint.xlsx", "store req. form")
        >>> print(cell_map.get('A1'))
    """
    try:
        wb = load_workbook(file_path, data_only=True)

        # Select sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.debug(f"Using sheet: {sheet_name}")
        else:
            ws = wb.worksheets[0]
            if sheet_name:
                logger.warning(
                    f"Sheet '{sheet_name}' not found, using first sheet: {ws.title}"
                )
            else:
                logger.debug(f"Using first sheet: {ws.title}")

        cell_map = {}

        for row in ws.iter_rows():
            for cell in row:
                if include_none or cell.value is not None:
                    cell_map[cell.coordinate] = cell.value

        logger.debug(f"Created cell map with {len(cell_map)} cells from {file_path}")

        return cell_map

    except Exception as e:
        logger.error(f"Error creating cell map from {file_path}: {e}", exc_info=True)
        raise


def clean_description(text: str | None) -> str | None:
    """
    Clean description text by normalizing whitespace.

    Args:
        text: Raw description text

    Returns:
        Cleaned text with normalized whitespace, or None if input is None

    Example:
        >>> clean_description("  Extra   spaces  \\n here  ")
        'Extra spaces here'
    """
    if not text:
        return text
    return re.sub(r"\s+", " ", str(text)).strip()


def parse_paint_description(text: str | None) -> dict[str, Any]:
    """
    Parse paint description to extract structured information.

    Extracts:
    - Chinese text
    - Product name
    - Color code (RAL or numeric)
    - Color name
    - Volume (e.g., "4L")

    Args:
        text: Raw description text

    Returns:
        Dictionary with keys: original, chinese, product_name, color_code,
        color_name, volume

    Example:
        >>> result = parse_paint_description("聚氨酯面漆 BLUE 4L RAL5001")
        >>> print(result['color_name'])  # 'BLUE'
        >>> print(result['volume'])  # '4L'
    """
    if not text:
        return {
            "original": None,
            "chinese": None,
            "product_name": None,
            "color_code": None,
            "color_name": None,
            "volume": None,
        }

    text = text.replace("，", " ").strip()

    # Split Chinese characters
    chinese = "".join(re.findall(r"[\u4e00-\u9fff]+", text))
    non_chinese = re.sub(r"[\u4e00-\u9fff]+", "", text).strip()

    tokens = non_chinese.split()

    # Extract volume (e.g., "4L")
    volume = None
    for t in tokens:
        if re.match(r"\d+L", t, re.IGNORECASE):
            volume = t
            break

    # Extract RAL code
    ral = next((t for t in tokens if t.upper().startswith("RAL")), None)

    # Extract numeric codes (3-5 digits)
    numeric_codes = [t for t in tokens if re.fullmatch(r"\d{3,5}", t)]

    # Detect special products (thinner, curing agent)
    is_special = any(keyword in chinese for keyword in CHINESE_SPECIAL_KEYWORDS)

    # Determine color code
    color_code = None
    if ral:
        color_code = ral
    elif numeric_codes and not is_special:
        color_code = numeric_codes[0]

    # Remove unwanted tokens
    cleaned_tokens = []
    for t in tokens:
        if t in {volume, color_code}:
            continue
        if t in IGNORE_TOKENS:
            continue
        cleaned_tokens.append(t)

    # Separate color tokens from product tokens
    color_tokens = []
    product_tokens = []

    for t in cleaned_tokens:
        if t.upper() in COLOR_LIST:
            color_tokens.append(t)
        else:
            product_tokens.append(t)

    # Handle multi-word colors (e.g., "BLUE GREEN")
    if len(color_tokens) > 1:
        color_name = " ".join(color_tokens)
    else:
        color_name = color_tokens[0] if color_tokens else None

    # Join product tokens
    product_name = " ".join(product_tokens).strip() if product_tokens else None

    return {
        "original": text,
        "chinese": chinese if chinese else None,
        "product_name": product_name,
        "color_code": color_code,
        "color_name": color_name,
        "volume": volume,
    }


def parse_store_req_form(
    cell_map: dict, column_mapping: dict[str, str] | None = None
) -> list[dict[str, Any]]:
    """
    Parse store request form data from cell map.

    Groups data by rows and extracts sections and items.

    Args:
        cell_map: Dictionary mapping cell coordinates to values
        column_mapping: Optional mapping of column letters to field names.
                       Defaults to standard ZO Paint mapping:
                       A: no, B: impa, C: description, D: unit,
                       E: rob, F: req, G: approved, H: remark

    Returns:
        List of dictionaries, one per item, with section information

    Example:
        >>> cell_map = create_cell_map("zo_paint.xlsx", "store req. form")
        >>> items = parse_store_req_form(cell_map)
        >>> for item in items:
        ...     print(f"{item['no']}: {item['description']}")
    """
    if column_mapping is None:
        column_mapping = {
            "A": "no",
            "B": "impa",
            "C": "description",
            "D": "unit",
            "E": "rob",
            "F": "req",
            "G": "approved",
            "H": "remark",
        }

    # Group by row
    rows = defaultdict(dict)
    for cell, val in cell_map.items():
        col = cell[0]
        row = int(cell[1:])
        rows[row][col] = val

    items = []
    current_section = None

    for row_idx in sorted(rows.keys()):
        row = rows[row_idx]

        # Detect section (has value in C but no value in A)
        if row.get("C") and not row.get("A"):
            current_section = str(row["C"]).strip()
            logger.debug(f"Detected section: {current_section}")
            continue

        # Detect item row (A is a number)
        if isinstance(row.get("A"), int | float):
            item = {"section": current_section}

            # Map columns to field names
            for col, field_name in column_mapping.items():
                item[field_name] = row.get(col)

            items.append(item)

    logger.info(f"Parsed {len(items)} items from store request form")

    return items


def parse_rfq_zo_paint(
    file_path: str, sheet_name: str = "store req. form", parse_descriptions: bool = True
) -> pd.DataFrame:
    """
    Parse ZO Paint RFQ Excel file and extract structured data.

    This function:
    1. Loads the Excel file and creates a cell map
    2. Parses the store request form
    3. Optionally parses paint descriptions for color/volume info
    4. Returns a cleaned DataFrame

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet containing the store request form
        parse_descriptions: Whether to parse descriptions for paint details

    Returns:
        DataFrame with RFQ data, including parsed paint info if enabled

    Example:
        >>> df = parse_rfq_zo_paint("zo_paint.xlsx")
        >>> print(df.columns)
        >>> print(df[['no', 'description', 'color_name', 'volume']].head())
    """
    # Create cell map
    cell_map = create_cell_map(file_path, sheet_name)

    # Parse store request form
    items = parse_store_req_form(cell_map)

    if not items:
        logger.warning(f"No items found in {file_path}")
        return pd.DataFrame()

    # Create DataFrame
    df = pd.DataFrame(items)

    # Parse descriptions if enabled
    if parse_descriptions and "description" in df.columns:
        df_clean = df["description"].apply(parse_paint_description).apply(pd.Series)
        df = pd.concat([df, df_clean], axis=1)
        logger.debug("Parsed paint descriptions for color/volume info")

    logger.info(f"Parsed ZO Paint RFQ: {len(df)} items from {file_path}")

    return df


def parse_rfq_zo_paint_to_dict(
    file_path: str, sheet_name: str = "store req. form", parse_descriptions: bool = True
) -> list[dict[str, Any]]:
    """
    Parse ZO Paint RFQ Excel file and return as list of dictionaries.

    This is a convenience function that converts the DataFrame result
    to a list of dicts for easier API integration.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet containing the store request form
        parse_descriptions: Whether to parse descriptions for paint details

    Returns:
        List of dictionaries, one per item in the parsed table

    Example:
        >>> items = parse_rfq_zo_paint_to_dict("zo_paint.xlsx")
        >>> for item in items:
        ...     print(f"{item['no']}: {item.get('color_name')}")
    """
    df = parse_rfq_zo_paint(
        file_path=file_path,
        sheet_name=sheet_name,
        parse_descriptions=parse_descriptions,
    )

    # Convert NaN to None for JSON serialization
    result = df.where(pd.notna(df), None).to_dict("records")

    return result


def batch_parse_zo_paint_excels(
    file_paths: list[str],
    sheet_name: str = "store req. form",
    parse_descriptions: bool = True,
) -> pd.DataFrame:
    """
    Parse multiple ZO Paint RFQ Excel files and combine into a single DataFrame.

    Args:
        file_paths: List of paths to Excel files
        sheet_name: Name of the sheet containing the store request form
        parse_descriptions: Whether to parse descriptions for paint details

    Returns:
        Combined DataFrame with data from all files

    Example:
        >>> files = ["zo1.xlsx", "zo2.xlsx", "zo3.xlsx"]
        >>> df = batch_parse_zo_paint_excels(files)
        >>> print(f"Total items: {len(df)}")
    """
    dataframes = []

    for file_path in file_paths:
        try:
            df = parse_rfq_zo_paint(
                file_path=file_path,
                sheet_name=sheet_name,
                parse_descriptions=parse_descriptions,
            )
            # Add source file column for tracking
            df["source_file"] = Path(file_path).name
            dataframes.append(df)

        except Exception as e:
            logger.error(f"Failed to parse {file_path}: {e}", exc_info=True)
            continue

    if not dataframes:
        logger.warning("No files were successfully parsed")
        return pd.DataFrame()

    # Combine all dataframes
    combined_df = pd.concat(dataframes, ignore_index=True)
    logger.info(f"Combined {len(dataframes)} files into {len(combined_df)} total items")

    return combined_df
