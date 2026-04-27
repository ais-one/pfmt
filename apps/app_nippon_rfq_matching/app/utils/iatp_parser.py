"""
IATP Excel Parser with multi-region and pricing support
Based on the original parser_excel_iatp_check_column_pricing logic
"""

import logging
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_excel_to_cell_mapping(file_path: str) -> dict[str, dict[str, str]]:
    """
    Parse Excel file to create cell address mapping

    Args:
        file_path: Path to Excel file

    Returns:
        Dictionary mapping sheet names to cell addresses and values
    """
    wb = load_workbook(file_path, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {}

        for row in ws.iter_rows():
            for cell in row:
                # Skip jika None / kosong
                if cell.value is None:
                    continue

                # Ambil alamat cell (contoh: A1, B2)
                cell_address = cell.coordinate

                # Simpan value
                sheet_data[cell_address] = str(cell.value)

        # Simpan per sheet
        result[sheet_name] = sheet_data

    return result


COLOR_COLUMN_MAP = {
    "IATP AF": 3,  # C
    "IATP GEN": 4,  # D
    "IATP SW": 5,  # E
}


def normalize_sheet_name(name):
    """Normalize sheet name by removing parentheses and converting to uppercase"""
    return name.upper().replace("(", "").replace(")", "").strip()


def get_color_column(sheet_name):
    """Get color column based on sheet name"""
    clean_name = normalize_sheet_name(sheet_name)

    for key, col in COLOR_COLUMN_MAP.items():
        if key in clean_name:
            return col

    return None  # No fallback - strict matching


def parse_excel_multi_sheet_flat(file_path: str) -> list[dict[str, Any]]:
    """
    Parse IATP Excel file with multi-region pricing support
    Uses strict color column matching without fallback

    Args:
        file_path: Path to the Excel file

    Returns:
        List of product entries with region pricing
    """
    wb = load_workbook(file_path, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.max_row < 7:
            continue

        # 🔥 GET COLOR COLUMN (NO GUESS)
        color_col = get_color_column(sheet_name)

        # --- STEP 1: Detect regions (row 4) ---
        region_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=4, column=col).value
            if val and isinstance(val, str):
                region_map[col] = val.strip()

        if not region_map:
            continue

        # --- STEP 2: Detect field type (row 5) ---
        field_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=5, column=col).value
            if val:
                v = str(val).lower()
                if "pack size" in v:
                    field_map[col] = "size"
                elif "uom" in v:
                    field_map[col] = "uom"
                elif "iatp" in v or "price" in v:
                    field_map[col] = "price"

        # --- STEP 3: Build region groups ---
        region_groups = {}
        current_region = None

        for col in range(1, ws.max_column + 1):
            if col in region_map:
                current_region = region_map[col]
                region_groups[current_region] = []

            if current_region and col in field_map:
                region_groups[current_region].append((col, field_map[col]))

        # --- STEP 4: Parse rows ---
        for row in range(7, ws.max_row + 1):
            product_code = ws.cell(row=row, column=1).value
            product_name = ws.cell(row=row, column=2).value

            if not product_code:
                continue

            # 🎨 COLOR (STRICT)
            color = None
            if color_col:
                val = ws.cell(row=row, column=color_col).value
                if val not in [None, "", "-"]:
                    color = str(val).strip()

            product_entry = {
                "product_code": str(product_code).strip(),
                "product_name": str(product_name).strip() if product_name else None,
                "color": color,
                "sheet_name": sheet_name,
                "regions": {},
            }

            for region, cols in region_groups.items():
                region_data = {}

                for col, field_type in cols:
                    value = ws.cell(row=row, column=col).value

                    if value in ["-", None, ""]:
                        continue

                    if field_type == "price":
                        price_str = str(value).strip().upper()
                        if price_str == "ENQUIRY":
                            region_data["price_raw"] = "Enquiry"
                            region_data["price"] = None
                        else:
                            try:
                                region_data["price"] = float(value)
                                region_data["price_raw"] = str(value)
                            except Exception:
                                region_data["price_raw"] = str(value)
                                region_data["price"] = None
                    else:
                        region_data[field_type] = value

                if region_data:
                    # Ensure price_raw exists even if price is None
                    if "price" in region_data and "price_raw" not in region_data:
                        region_data["price_raw"] = (
                            str(region_data["price"])
                            if region_data["price"] is not None
                            else None
                        )
                    product_entry["regions"][region] = region_data

            results.append(product_entry)

    return results


def parse_excel_to_region_json(
    file_path: str, sheet_name: str = None
) -> dict[str, Any]:
    """
    Parse IATP Excel file with multi-region support and pricing data
    Adapted from original parser logic

    Args:
        file_path: Path to Excel file
        sheet_name: Optional specific sheet to parse (if None, process all sheets)

    Returns:
        Dictionary with products, regions, and summary
    """
    wb = load_workbook(file_path, data_only=True)

    # If no specific sheet name provided, process all sheets
    if sheet_name is None:
        sheet_names = wb.sheetnames
    else:
        sheet_names = [sheet_name]

    all_products = []
    all_regions = set()

    # Process each sheet
    for sheet in sheet_names:
        ws = wb[sheet]
        logger.info(f"Processing sheet: {sheet}")

        # --- STEP 1: Detect regions (row 4) ---
        region_map = {}  # col_index -> region_name
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=4, column=col).value
            if val and isinstance(val, str):
                region_map[col] = val.strip()
                all_regions.add(val.strip())

        # --- STEP 2: Detect field type (row 5) ---
        field_map = {}  # col_index -> field_type
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=5, column=col).value
            if val:
                v = str(val).lower()
                if "pack size" in v:
                    field_map[col] = "size"
                elif "uom" in v:
                    field_map[col] = "uom"
                elif "iatp" in v or "price" in v:
                    field_map[col] = "price"

        # --- STEP 3: Build region column groups ---
        region_groups = {}
        current_region = None

        for col in range(1, ws.max_column + 1):
            if col in region_map:
                current_region = region_map[col]
                region_groups[current_region] = []

            if current_region and col in field_map:
                region_groups[current_region].append((col, field_map[col]))

        # --- STEP 4: Parse rows (data starts from row 6 or 7 depending on sheet type) ---
        products = []

        # Determine start row based on sheet type
        if "SW" in sheet:
            start_row = 6  # SW starts from row 6
        else:
            start_row = 7  # AF and GEN start from row 7

        logger.info(f"Starting to parse rows from row {start_row} for sheet: {sheet}")

        for row in range(start_row, ws.max_row + 1):
            product_code = ws.cell(row=row, column=1).value
            product_name = ws.cell(row=row, column=2).value

            if not product_code:
                continue

            # Extract color based on sheet type
            color = None
            if "SW" in sheet:
                # SW: color is in column E (index 5)
                color = ws.cell(row=row, column=5).value
            elif "AF" in sheet:
                # AF: color is in column C (index 3), fallback to D (index 4)
                color = ws.cell(row=row, column=3).value
                if not color:
                    color = ws.cell(row=row, column=4).value
            elif "GEN" in sheet:
                # GEN: color is in column D (index 4)
                color = ws.cell(row=row, column=4).value

            # Clean up color value
            if isinstance(color, str) and color.strip() == "":
                color = None

            product_entry = {
                "product_code": product_code,
                "product_name": product_name,
                "sheet_name": sheet,
                "sheet_type": "IATP_MULTI_REGION",
                "row_excel": row,
                "pmc": product_code,
                "color": color,
                "region_prices": {},
            }

            # Parse pricing data for each region
            for region, cols in region_groups.items():
                region_data = {}

                for col, field_type in cols:
                    value = ws.cell(row=row, column=col).value

                    if value in ["-", None, ""]:
                        continue

                    region_data[field_type] = value

                # Only add if we have pricing data
                if region_data and "price" in region_data:
                    # Convert price to float if possible, store raw value
                    price_clean = None
                    price_raw = None
                    price = region_data["price"]

                    price_str = str(price).strip().upper()
                    if price_str == "ENQUIRY":
                        price_raw = "Enquiry"
                    elif price_str and price_str not in ["-", ""]:
                        try:
                            price_clean = float(price_str)
                        except Exception:
                            price_raw = price_str

                    if price_clean is not None or price_raw == "Enquiry":
                        pricing_record = {
                            "pmc": product_code,
                            "product_name": product_name,
                            "region": region,
                            "size": float(region_data["size"])
                            if "size" in region_data
                            and str(region_data["size"])
                            .replace(".", "")
                            .replace(",", "")
                            .isdigit()
                            else None,
                            "uom": str(region_data["uom"])
                            if "uom" in region_data
                            else None,
                            "price": price_clean,
                            "price_raw": price_raw,
                        }

                        product_entry["region_prices"][region] = pricing_record

            all_products.extend(products)

    # Remove duplicate products based on product_code and sheet_name
    unique_products = []
    seen = set()

    for product in all_products:
        key = (product["product_code"], product["sheet_name"])
        if key not in seen:
            seen.add(key)
            unique_products.append(product)

    all_products = unique_products

    # Generate summary
    summary = {
        "total_products": len(all_products),
        "total_regions": len(all_regions),
        "regions": list(all_regions),
        "products_with_pricing": sum(1 for p in all_products if p["region_prices"]),
        "parse_timestamp": datetime.now().isoformat(),
    }

    return {"products": all_products, "regions": list(all_regions), "summary": summary}


def parse_iatp_excel_with_multi_region_and_pricing(file_path: str) -> dict[str, Any]:
    """
    Parse IATP Excel file with multi-region pricing support
    Uses the improved logic with strict color column matching

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary containing parsed data
    """
    products = parse_excel_multi_sheet_flat(file_path)

    # Extract all unique regions
    all_regions = set()
    for product in products:
        if "regions" in product:
            for region in product["regions"].keys():
                all_regions.add(region)

    # Generate summary
    summary = {
        "total_products": len(products),
        "total_regions": len(all_regions),
        "regions": list(all_regions),
        "products_with_pricing": sum(1 for p in products if p.get("regions")),
        "parse_timestamp": datetime.now().isoformat(),
    }

    return {"products": products, "regions": list(all_regions), "summary": summary}
