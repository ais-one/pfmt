"""
Utility functions for parsing IATP Excel files and RFQ PDF files
"""

import logging
import re
import uuid
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd

from apps.app_nippon_rfq_matching.app.utils.iatp_parser import (
    parse_iatp_excel_with_multi_region_and_pricing,
)


def generate_code() -> str:
    """Generate unique 8-character code"""
    return str(uuid.uuid4())[:8]


def parse_iatp_af(df: pd.DataFrame, sheet_name: str) -> list[dict[str, Any]]:
    """
    Parse IATP (AF) sheet type

    Args:
        df: DataFrame from Excel sheet
        sheet_name: Name of the sheet

    Returns:
        List of parsed product records
    """
    results = []
    df = df.iloc[6:]  # Start from row 7

    for idx, row in df.iterrows():
        pmc = row[0]  # Column A
        product_name = row[1]  # Column B
        color = row[2]  # Column C

        # Fallback color to column D if C is empty
        if pd.isna(color):
            color = row[3]

        if pd.notna(product_name):
            results.append(
                {
                    "sheet_name": sheet_name,
                    "sheet_type": "IATP_AF",
                    "row_excel": idx + 1,
                    "pmc": pmc,
                    "product_name": product_name,
                    "color": color,
                }
            )

    return results


def parse_iatp_sw(df: pd.DataFrame, sheet_name: str) -> list[dict[str, Any]]:
    """
    Parse IATP (SW) sheet type

    Args:
        df: DataFrame from Excel sheet
        sheet_name: Name of the sheet

    Returns:
        List of parsed product records
    """
    results = []
    df = df.iloc[5:]  # Start from row 6

    for idx, row in df.iterrows():
        product_name = row[1]  # Column B
        color = row[4]  # Column E

        # Treat empty string as None
        if isinstance(product_name, str) and product_name.strip() == "":
            product_name = None

        if isinstance(color, str) and color.strip() == "":
            color = None

        # Skip if both B and E are empty
        if pd.isna(product_name) and pd.isna(color):
            continue

        results.append(
            {
                "sheet_name": sheet_name,
                "sheet_type": "IATP_SW",
                "row_excel": idx + 1,
                "pmc": generate_code(),
                "product_name": product_name,
                "color": color,
            }
        )

    return results


def parse_iatp_gen(df: pd.DataFrame, sheet_name: str) -> list[dict[str, Any]]:
    """
    Parse IATP (GEN) sheet type

    Args:
        df: DataFrame from Excel sheet
        sheet_name: Name of the sheet

    Returns:
        List of parsed product records
    """
    results = []
    df = df.iloc[6:]  # Start from row 7

    for idx, row in df.iterrows():
        pmc = row[0]  # Column A
        product_name = row[1]  # Column B
        color = row[3]  # Column D

        if isinstance(product_name, str) and product_name.strip() == "":
            product_name = None

        if isinstance(color, str) and color.strip() == "":
            color = None

        # Skip if both B and D are empty
        if pd.isna(product_name) and pd.isna(color):
            continue

        results.append(
            {
                "sheet_name": sheet_name,
                "sheet_type": "IATP_GEN",
                "row_excel": idx + 1,
                "pmc": pmc,
                "product_name": product_name,
                "color": color,
            }
        )

    return results


def parse_iatp_excel(file_path: str) -> list[dict[str, Any]]:
    """
    Parse IATP Excel file and extract all product data
    Uses the improved multi-region parser that handles all sheet types automatically

    Args:
        file_path: Path to the Excel file

    Returns:
        List of all parsed product records
    """
    # Use the improved multi-region parser
    data = parse_iatp_excel_with_multi_region_and_pricing(file_path)

    # Convert the format to match expected return structure
    # Filter out products without color to match old behavior
    filtered_results = [
        {
            "sheet_name": product.get("sheet_name"),
            "sheet_type": product.get("sheet_type"),
            "row_excel": product.get("row_excel"),
            "pmc": product.get("product_code"),
            "product_name": product.get("product_name"),
            "color": product.get("color"),
        }
        for product in data["products"]
        if product.get("color")
    ]

    return filtered_results


def parse_iatp_excel_with_multi_region(file_path: str) -> dict[str, Any]:
    """
    Parse IATP Excel file with multi-region pricing support
    Now uses the combined parser function

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary containing:
        - 'products': List of product records with pricing data
        - 'regions': List of detected regions
        - 'summary': Processing summary
    """
    return parse_iatp_excel_with_multi_region_and_pricing(file_path)


def parse_iatp_excel_with_pricing(file_path: str) -> dict[str, list[dict[str, Any]]]:
    """
    Parse IATP Excel file with pricing support
    Now uses the combined parser function

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary containing:
        - 'products': List of product records
        - 'pricing': List of pricing records
    """
    data = parse_iatp_excel_with_multi_region_and_pricing(file_path)

    # Convert format to match expected return structure
    return {
        "products": data["products"],
        "pricing": [
            price
            for product in data["products"]
            for price in product["regions"].values()
        ],
    }


def clean_column(col: str) -> str:
    """Clean column name by extracting last part after dot"""
    return col.split(".")[-1]


def find_column(columns: list[str], target: str) -> str | None:
    """
    Find a column by partial name match.
    Handles merged column names like '# Part Type' when looking for '#'

    Args:
        columns: List of column names
        target: Target column name to find

    Returns:
        Matching column name or None
    """
    import logging

    logger = logging.getLogger(__name__)

    # First try exact match
    if target in columns:
        logger.debug(f"  find_column: Exact match found for '{target}'")
        return target

    # Try partial match (for merged columns like '# Part Type' containing '#')
    for col in columns:
        if target in col:
            logger.debug(
                f"  find_column: Partial match found - '{col}' contains '{target}'"
            )
            return col

    logger.warning(f"  find_column: No match found for '{target}' in columns {columns}")
    return None


def parse_rfq1(rfq: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Parse RFQ1 type (from PDF table extraction)

    Args:
        rfq: Dictionary containing table data

    Returns:
        List of parsed RFQ items
    """
    import logging

    logger = logging.getLogger(__name__)

    results = []

    # Table 0
    t0 = rfq.get("table_0", {})
    if t0:
        try:
            item = {v[0]: v[1] for v in t0.values()}
            results.append(
                {
                    "raw_text": item.get("Description", ""),
                    "qty": item.get("Qty"),
                    "uom": item.get("UoM"),
                    "source": "rfq_1_table_0",
                }
            )
        except Exception as e:
            logger.warning(f"Failed to parse table_0: {e}")

    # Table 1
    t1 = rfq.get("table_1", {})
    if t1:
        try:
            logger.info(
                f"parse_rfq1: Processing table_1 with columns: {list(t1.keys())}"
            )

            # Find columns by partial match (handles merged column names)
            hash_col = find_column(list(t1.keys()), "#")
            supplier_col = find_column(list(t1.keys()), "Supplier Part No.")
            desc_col = find_column(list(t1.keys()), "Description")
            uom_col = find_column(list(t1.keys()), "UoM")
            qty_col = find_column(list(t1.keys()), "Qty")

            logger.info(
                f"parse_rfq1: Found columns - #: {hash_col}, Supplier: {supplier_col}, Desc: {desc_col}, "
                f"UoM: {uom_col}, Qty: {qty_col}"
            )

            if hash_col:
                for i in range(len(t1[hash_col])):
                    text = t1.get(supplier_col, {}).get(i, "") or ""

                    # Fallback to description if supplier part no is empty
                    if not text and desc_col:
                        text = t1.get(desc_col, {}).get(i, "")

                    results.append(
                        {
                            "raw_text": text,
                            "qty": t1.get(qty_col, {}).get(i) if qty_col else None,
                            "uom": t1.get(uom_col, {}).get(i) if uom_col else None,
                            "source": "rfq_1_table_1",
                        }
                    )
            else:
                logger.warning("parse_rfq1: Could not find '#' column in table_1")
        except Exception as e:
            logger.error(f"Error parsing table_1: {e}", exc_info=True)

    return results


def parse_rfq2(rfq: dict[str, Any]) -> list[dict[str, Any]]:
    """
    Parse RFQ2 type (from PDF table extraction)

    Args:
        rfq: Dictionary containing table data

    Returns:
        List of parsed RFQ items
    """
    results = []

    t0 = rfq.get("table_0", {})
    if t0:
        # Find columns by partial match (handles merged column names)
        hash_col = find_column(list(t0.keys()), "#")
        desc_col = find_column(list(t0.keys()), "Description")
        uom_col = find_column(list(t0.keys()), "UoM")
        qty_col = find_column(list(t0.keys()), "Qty")

        if hash_col:
            n = len(t0[hash_col])
            for i in range(n):
                results.append(
                    {
                        "raw_text": t0.get(desc_col, {}).get(i, "") if desc_col else "",
                        "qty": t0.get(qty_col, {}).get(i) if qty_col else None,
                        "uom": t0.get(uom_col, {}).get(i) if uom_col else None,
                        "source": "rfq_2",
                    }
                )

    return results


def parse_all_rfq(tables_dict: dict[str, pd.DataFrame]) -> list[dict[str, Any]]:
    """
    Parse all RFQ items from RFQ1 and RFQ2

    Args:
        tables_dict: Dictionary containing table data

    Returns:
        List of all parsed RFQ items
    """
    all_items = []

    all_items.extend(parse_rfq1(tables_dict))

    all_items.extend(parse_rfq2(tables_dict))

    return all_items


def clean_raw_text_rfq(text: str) -> str:
    """
    Clean raw text from RFQ for matching

    Args:
        text: Raw text string

    Returns:
        Cleaned text string
    """
    text = text.upper()
    text = re.sub(r"\[.*?\]", "", text)
    text = re.sub(r"\(.*?\)", "", text)
    text = re.sub(r"-UOM:.*", "", text)
    text = re.sub(r"[^A-Z0-9 ]", " ", text)
    return " ".join(text.split())


def parse_ceru_excel(file_path: str) -> list[dict[str, Any]]:
    """
    Parse CERU Excel file format (RFQ with prefix CERU-)

    Format specification:
    - Sheet: 'REQUISITION' or first sheet
    - Header columns: B20:M20 (row 20, columns B to M)
    - Data rows: B22:M26 (rows 22-26, columns B to M)

    Args:
        file_path: Path to the Excel file

    Returns:
        List of parsed RFQ items with keys: raw_text, qty, uom, source

    Example:
        >>> items = parse_ceru_excel("CERU-D-25-096.xlsx")
        >>> print(items[0])
        {'raw_text': 'NIPPON MARINE THINNER 200', 'qty': '120', 'uom': 'LTR', 'source': 'ceru_excel'}
    """
    import logging

    logger = logging.getLogger(__name__)

    rfq_items = []

    try:
        # Read Excel file
        xls = pd.ExcelFile(file_path)

        # Try to find 'REQUISITION' sheet, otherwise use first sheet
        sheet_name = None
        if "REQUISITION" in xls.sheet_names:
            sheet_name = "REQUISITION"
            logger.info(f"Found 'REQUISITION' sheet in {file_path}")
        else:
            sheet_name = xls.sheet_names[0]
            logger.info(f"Using first sheet '{sheet_name}' in {file_path}")

        # Read the sheet without header (header=None)
        # Read all data to get the structure
        df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        logger.info(f"Sheet shape: {df_full.shape}")
        logger.debug(f"Full sheet preview:\n{df_full.head(30)}")

        # Define header and data row positions
        # Header is at row 20 (index 19 in 0-based)
        # Data is at rows 22-26 (indices 21-25 in 0-based)
        # Columns are B to M (indices 1-12 in 0-based, where A=0, B=1, ..., M=12)

        header_row_idx = 19  # Row 20 (0-based)
        data_start_row = 21  # Row 22 (0-based)
        data_end_row = 25  # Row 26 (0-based)
        col_start = 1  # Column B (0-based)
        col_end = 12  # Column M (0-based)

        # Check if we have enough rows
        if len(df_full) <= header_row_idx:
            logger.error(
                f"Sheet has only {len(df_full)} rows, need at least {header_row_idx + 1} rows"
            )
            return []

        # Extract header row
        headers = df_full.iloc[header_row_idx, col_start : col_end + 1].tolist()
        logger.info(f"Headers found: {headers}")

        # Extract data rows
        for row_idx in range(data_start_row, min(data_end_row + 1, len(df_full))):
            row_data = df_full.iloc[row_idx, col_start : col_end + 1].tolist()

            # Skip empty rows
            if all(pd.isna(val) or val == "" for val in row_data):
                logger.debug(f"Skipping empty row {row_idx}")
                continue

            # Create dict from headers and data
            row_dict = dict(zip(headers, row_data))

            logger.debug(f"Row {row_idx} data: {row_dict}")

            # Extract RFQ item data
            # Look for common column names
            raw_text = None
            qty = None
            uom = None

            # Try to find description/material column
            for key in [
                "Material",
                "Description",
                "Item",
                "Product",
                "Material Description",
            ]:
                if key in row_dict and pd.notna(row_dict[key]):
                    raw_text = str(row_dict[key]).strip()
                    break

            # Try to find quantity column
            for key in ["Quantity", "Qty", "Order Quantity", "Req. Qty"]:
                if key in row_dict and pd.notna(row_dict[key]):
                    qty_val = row_dict[key]
                    # Convert to string, handle numeric values
                    if isinstance(qty_val, int | float):
                        qty = (
                            str(int(qty_val))
                            if qty_val == int(qty_val)
                            else str(qty_val)
                        )
                    else:
                        qty = str(qty_val).strip()
                    break

            # Try to find UOM column
            for key in ["UoM", "UOM", "Unit", "Unit of Measure"]:
                if key in row_dict and pd.notna(row_dict[key]):
                    uom = str(row_dict[key]).strip()
                    break

            # If not found in named columns, try positional mapping
            if not raw_text:
                # Usually description is in the first or second column
                for i, val in enumerate(row_data):
                    if pd.notna(val) and isinstance(val, str) and len(val.strip()) > 3:
                        raw_text = val.strip()
                        break

            if not qty:
                # Usually quantity is in a later column
                for i, val in enumerate(row_data):
                    if pd.notna(val) and isinstance(val, int | float):
                        qty = str(int(val)) if val == int(val) else str(val)
                        break

            if not uom:
                # Usually UOM is after quantity
                for i, val in enumerate(row_data):
                    if (
                        pd.notna(val)
                        and isinstance(val, str)
                        and val.strip().upper()
                        in ["LTR", "KG", "PCS", "EA", "SET", "M", "MT"]
                    ):
                        uom = val.strip()
                        break

            # Create RFQ item if we have at least raw_text
            if raw_text:
                rfq_item = {
                    "raw_text": raw_text,
                    "qty": qty,
                    "uom": uom,
                    "source": "ceru_excel",
                }
                rfq_items.append(rfq_item)
                logger.debug(f"Extracted RFQ item: {rfq_item}")
            else:
                logger.warning(
                    f"Could not extract raw_text from row {row_idx}: {row_dict}"
                )

        logger.info(f"Parsed {len(rfq_items)} RFQ items from CERU Excel file")

        # Add clean text
        for item in rfq_items:
            if item.get("raw_text"):
                item["clean_text"] = clean_raw_text_rfq(item["raw_text"])

    except Exception as e:
        logger.error(f"Error parsing CERU Excel file {file_path}: {e}", exc_info=True)
        raise

    return rfq_items


def is_ceru_file(filename: str) -> bool:
    """
    Check if filename matches CERU format

    Args:
        filename: Filename to check

    Returns:
        True if filename starts with CERU (case-insensitive)
    """
    return filename.upper().startswith("CERU-")


def insert_regions_to_database(regions: list[str], db) -> dict[str, int]:
    """
    Insert regions into database if they don't exist

    Args:
        regions: List of region names
        db: Database session

    Returns:
        Dictionary mapping region names to their IDs
    """
    from apps.app_nippon_rfq_matching.app.models.pricing import Region

    # Define logger
    logger = logging.getLogger(__name__)

    region_map = {}

    for region_name in regions:
        # Check if region exists
        region = db.query(Region).filter(Region.name == region_name).first()

        if not region:
            # Create new region
            region = Region(name=region_name)
            db.add(region)

        region_map[region_name] = region

    # Commit all new regions
    try:
        db.commit()
        # Refresh to get IDs
        for region_name in region_map:
            region = db.query(Region).filter(Region.name == region_name).first()
            region_map[region_name] = region.id
    except Exception as e:
        db.rollback()
        logger.error(f"Error inserting regions: {str(e)}")
        raise

    return region_map


def insert_product_prices_to_database(
    products_data: list[dict[str, Any]], region_map: dict[str, int], db
) -> int:
    """
    Insert product prices into database

    Args:
        products_data: List of product data with region_prices
        region_map: Mapping of region names to region IDs
        db: Database session

    Returns:
        Number of records inserted
    """
    from apps.app_nippon_rfq_matching.app.models.database import ProductMaster
    from apps.app_nippon_rfq_matching.app.models.pricing import ProductPrices

    # Define logger
    logger = logging.getLogger(__name__)

    inserted_count = 0

    for product_data in products_data:
        product_code = product_data.get("product_code")
        region_prices = product_data.get("regions", {})

        # Debug: Check regions format
        if not isinstance(region_prices, dict):
            logger.error(
                f"region_prices is not a dict for product {product_code}: {region_prices} (type: {type(region_prices)})"
            )
            continue

        # Find product master
        product_master = (
            db.query(ProductMaster).filter(ProductMaster.pmc == product_code).first()
        )

        if not product_master:
            logger.warning(f"Product master not found for code: {product_code}")
            continue

        # Insert pricing for each region
        for region_name, price_data in region_prices.items():
            if region_name not in region_map:
                logger.warning(f"Region not found: {region_name}")
                continue

            # Skip if no valid price data
            if not price_data.get("price") and price_data.get("price_raw") != "Enquiry":
                logger.debug(f"Skipping region {region_name} due to missing price data")
                continue

            # Check if pricing already exists
            existing_price = (
                db.query(ProductPrices)
                .filter(
                    ProductPrices.product_master_id == product_master.id,
                    ProductPrices.region_id == region_map[region_name],
                    ProductPrices.size == price_data.get("size"),
                    ProductPrices.uom == price_data.get("uom"),
                )
                .first()
            )

            if existing_price:
                logger.info(
                    f"Pricing already exists for product {product_code} in {region_name}"
                )
                continue

            # Create new pricing record
            pricing_record = ProductPrices(
                product_master_id=product_master.id,
                region_id=region_map[region_name],
                size=price_data.get("size"),
                uom=price_data.get("uom"),
                price=price_data.get("price"),
                price_raw=price_data.get("price_raw"),
                created_at=datetime.now(),
            )

            db.add(pricing_record)
            inserted_count += 1

    try:
        db.commit()
        logger.info(f"Successfully inserted {inserted_count} pricing records")
    except Exception as e:
        db.rollback()
        logger.error(f"Error inserting product prices: {str(e)}")
        raise

    return inserted_count


def process_iatp_excel_with_database_insertion(file_path: str, db) -> dict[str, Any]:
    """
    Complete pipeline: Parse Excel with multi-region support and insert data into database

    Args:
        file_path: Path to Excel file
        db: Database session

    Returns:
        Dictionary containing processing results and statistics
    """
    from apps.app_nippon_rfq_matching.app.models.database import ProductMaster

    # Define logger
    logger = logging.getLogger(__name__)

    results = {
        "status": "success",
        "products_inserted": 0,
        "regions_inserted": 0,
        "pricing_records_inserted": 0,
        "errors": [],
        "timestamp": datetime.now().isoformat(),
    }

    try:
        # Step 1: Parse the Excel file with multi-region support
        logger.info(
            f"Starting to parse Excel file with multi-region support: {file_path}"
        )
        parsed_data = parse_iatp_excel_with_multi_region_and_pricing(file_path)
        products = parsed_data.get("products", [])

        # Extract all unique regions from products
        all_regions = set()
        for product in products:
            for region in product.get("regions", {}).keys():
                all_regions.add(region)

        # Step 2: Insert regions
        logger.info("Inserting regions into database")
        region_map = insert_regions_to_database(list(all_regions), db)
        results["regions_inserted"] = len(region_map)

        # Step 3: Create product master records with conflict resolution
        logger.info("Creating product master records with conflict resolution")
        products_inserted = 0
        products_updated = 0

        for product_data in products:
            product_code = product_data.get("product_code")
            product_name = product_data.get("product_name")
            color = product_data.get("color")
            sheet_name = product_data.get("sheet_name")
            sheet_type = product_data.get("sheet_type", "IATP_MULTI_REGION")

            # Check if product already exists
            existing_product = (
                db.query(ProductMaster)
                .filter(
                    ProductMaster.sheet_name == sheet_name,
                    ProductMaster.pmc == product_code,
                    ProductMaster.product_name == product_name,
                    ProductMaster.color == color,
                )
                .first()
            )

            if existing_product:
                # Product exists, update if needed (for multi-region support)
                try:
                    updated = False
                    if existing_product.sheet_type != sheet_type:
                        existing_product.sheet_type = sheet_type
                        updated = True
                    if existing_product.product_name != product_name:
                        existing_product.product_name = product_name
                        existing_product.clean_product_name = (
                            clean_raw_text_rfq(product_name) if product_name else None
                        )
                        updated = True
                    if existing_product.color != color:
                        existing_product.color = color
                        updated = True

                    if updated:
                        products_updated += 1
                        logger.debug(
                            f"Updated product master: {product_code} - {product_name} (color: {color}, "
                            f"sheet: {sheet_type})"
                        )

                except Exception as e:
                    logger.error(f"Error updating product {product_code}: {str(e)}")
                    continue
            else:
                # Create new product master
                try:
                    product_master = ProductMaster(
                        sheet_name=sheet_name,
                        sheet_type=sheet_type,
                        row_excel=None,  # Not available in new format
                        pmc=product_code,
                        product_name=product_name,
                        color=color,
                        clean_product_name=clean_raw_text_rfq(product_name)
                        if product_name
                        else None,
                        uploaded_file_id=None,  # Will be set later
                    )
                    db.add(product_master)
                    products_inserted += 1
                    logger.debug(
                        f"Added product master: {product_code} - {product_name} (color: {color}, sheet: {sheet_type})"
                    )

                except Exception as e:
                    # If unique constraint violation (race condition), skip this product
                    if "UNIQUE constraint failed" in str(e) or "IntegrityError" in str(
                        type(e).__name__
                    ):
                        logger.debug(
                            f"Product already exists, skipping: {product_code} - {product_name} (color: {color})"
                        )
                        continue
                    else:
                        # Re-raise other errors
                        logger.error(
                            f"Error inserting product {product_code}: {str(e)}"
                        )
                        raise

        # Commit product master records
        try:
            db.commit()
            logger.info(
                f"Successfully processed {products_inserted + products_updated} product master records "
                f"({products_inserted} inserted, {products_updated} updated)"
            )
            results["products_inserted"] = products_inserted
            results["products_updated"] = products_updated
        except Exception as e:
            db.rollback()
            logger.error(f"Error creating product masters: {str(e)}")
            raise

        # Step 4: Insert product prices
        logger.info("Inserting product prices into database")

        # Debug: Check format of products data
        if products:
            sample_product = products[0]
            logger.debug(f"Sample product format: {sample_product.keys()}")
            if "regions" in sample_product:
                logger.debug(f"Sample regions: {sample_product['regions']}")
            else:
                logger.error("No 'regions' key found in product data!")

        try:
            logger.info(
                f"About to call insert_product_prices_to_database with {len(products)} products"
            )
            # Log first product details
            if products:
                logger.info(f"First product: {products[0]}")
            pricing_count = insert_product_prices_to_database(products, region_map, db)
            results["pricing_records_inserted"] = pricing_count
        except Exception as e:
            logger.error(f"Error inserting prices: {str(e)}")
            logger.error(f"Error type: {type(e).__name__}")
            import traceback

            traceback.print_exc()
            # Log more details about the products data
            if products:
                logger.error(f"First product: {products[0]}")
            raise

        # Update statistics
        results["products_inserted"] = len(products)
        if "summary" in parsed_data:
            results["summary"] = parsed_data["summary"]

    except Exception as e:
        results["status"] = "error"
        results["errors"].append(str(e))
        logger.error(f"Error processing Excel file with database insertion: {str(e)}")
        raise

    return results


COLOR_COLUMN_MAP = {
    "IATP AF": 3,  # C
    "IATP GEN": 4,  # D
    "IATP SW": 5,  # E
}


def normalize_sheet_name(name):
    return name.upper().replace("(", "").replace(")", "").strip()


def get_color_column(sheet_name):
    clean_name = normalize_sheet_name(sheet_name)

    for key, col in COLOR_COLUMN_MAP.items():
        if key in clean_name:
            return col

    return None  # ❗ tidak pakai fallback


def parse_excel_multi_sheet_flat(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if ws.max_row < 7:
            continue

        # 🔥 GET COLOR COLUMN (NO GUESS)
        color_col = get_color_column(sheet_name)

        # --- STEP 1: Detect regions (row 4) ---
        region_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=4, column=col).value
            if val and isinstance(val, str):
                region_map[col] = val.strip()

        if not region_map:
            continue

        # --- STEP 2: Detect field type (row 5) ---
        field_map = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=5, column=col).value
            if val:
                v = str(val).lower()
                if "pack size" in v:
                    field_map[col] = "size"
                elif "uom" in v:
                    field_map[col] = "uom"
                elif "iatp" in v or "price" in v:
                    field_map[col] = "price"

        # --- STEP 3: Build region groups ---
        region_groups = {}
        current_region = None

        for col in range(1, ws.max_column + 1):
            if col in region_map:
                current_region = region_map[col]
                region_groups[current_region] = []

            if current_region and col in field_map:
                region_groups[current_region].append((col, field_map[col]))

        # --- STEP 4: Parse rows ---
        for row in range(7, ws.max_row + 1):
            product_code = ws.cell(row=row, column=1).value
            product_name = ws.cell(row=row, column=2).value

            if not product_code:
                continue

            # 🔥 COLOR (STRICT)
            color = None
            if color_col:
                val = ws.cell(row=row, column=color_col).value
                if val not in [None, "", "-"]:
                    color = str(val).strip()

            product_entry = {
                "product_code": str(product_code).strip(),
                "product_name": str(product_name).strip() if product_name else None,
                "color": color,
                "sheet_name": sheet_name,
                "regions": {},
            }

            for region, cols in region_groups.items():
                region_data = {}

                for col, field_type in cols:
                    value = ws.cell(row=row, column=col).value

                    if value in ["-", None, ""]:
                        continue

                    if field_type == "price":
                        try:
                            value = float(value)
                        except Exception:
                            value = str(value)

                    region_data[field_type] = value

                if region_data:
                    product_entry["regions"][region] = region_data

            results.append(product_entry)

    return results
