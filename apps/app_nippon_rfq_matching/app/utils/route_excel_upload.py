"""Excel upload router.

Routes Excel RFQ uploads to the appropriate parser based on format detection.
Supports CERU, ZO Paint, and Indonesian RFQ formats.
"""

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from apps.app_nippon_rfq_matching.app.utils.rfq_excel_ceru_parser import (
    is_ceru_excel,
    parse_ceru_excel,
)
from apps.app_nippon_rfq_matching.app.utils.rfq_indonesia_parser import (
    is_indonesia_rfqs,
    parse_rfq_excel_to_dict,
)
from apps.app_nippon_rfq_matching.app.utils.rfq_zo_paint_parser import (
    is_zo_paint_excel,
    parse_rfq_zo_paint_to_dict,
)

logger = logging.getLogger(__name__)

# Parser type constants
PARSER_CERU = "ceru"
PARSER_ZO_PAINT = "zo_paint"
PARSER_INDONESIA = "indonesia"
PARSER_UNKNOWN = "unknown"


def create_cell_map(file_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """Create a mapping of cell coordinates to their values.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to load. If None, uses the first sheet.

    Returns:
        Dictionary mapping cell coordinates (e.g., 'A1') to their values.
    """
    try:
        wb = load_workbook(file_path, data_only=True)

        # Select sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[0]

        cell_map = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_map[cell.coordinate] = cell.value

        logger.debug(f"Created cell map with {len(cell_map)} cells from {file_path}")

        return cell_map

    except Exception as e:
        logger.error(f"Error creating cell map from {file_path}: {e}", exc_info=True)
        raise


def detect_parser_type(cell_map: dict[str, Any], sheet_name: str | None = None) -> str:
    """Detect the appropriate parser type for the given cell data.

    Args:
        cell_map: Dictionary mapping cell coordinates to values.
        sheet_name: Name of the sheet (helps with identification).

    Returns:
        Parser type constant: PARSER_CERU, PARSER_ZO_PAINT, PARSER_INDONESIA,
        or PARSER_UNKNOWN.
    """
    # Check in priority order
    if is_ceru_excel(cell_map):
        return PARSER_CERU

    if is_zo_paint_excel(cell_map, sheet_name):
        return PARSER_ZO_PAINT

    if is_indonesia_rfqs(cell_map, sheet_name):
        return PARSER_INDONESIA

    return PARSER_UNKNOWN


def route_and_parse(
    file_path: str, sheet_name: str | None = None, parser_type: str | None = None
) -> dict[str, Any]:
    """Route Excel file to appropriate parser and parse the data.

    Args:
        file_path: Path to the Excel file.
        sheet_name: Name of the sheet to parse (optional).
        parser_type: Force specific parser type. If None, auto-detects.
                     Options: "ceru", "zo_paint", "indonesia"

    Returns:
        Dictionary with keys:
        - "parser_type": The parser used ("ceru", "zo_paint", "indonesia", "unknown")
        - "data": Parsed data (format varies by parser)
        - "error": Error message if parsing failed (optional)

    Raises:
        ValueError: If file doesn't exist or parser_type is invalid.
        Exception: If parsing fails.
    """
    file_path_obj = Path(file_path)

    if not file_path_obj.exists():
        raise ValueError(f"File not found: {file_path}")

    # Create cell map for detection
    cell_map = create_cell_map(file_path, sheet_name)

    # Auto-detect parser type if not specified
    if parser_type is None:
        detected_type = detect_parser_type(cell_map, sheet_name)
        parser_type = detected_type
        logger.info(f"Auto-detected parser type: {parser_type}")
    else:
        # Validate parser type
        valid_types = {PARSER_CERU, PARSER_ZO_PAINT, PARSER_INDONESIA}
        if parser_type not in valid_types:
            raise ValueError(
                f"Invalid parser_type: {parser_type}. Must be one of {valid_types}"
            )

    # Route to appropriate parser
    try:
        if parser_type == PARSER_CERU:
            data = parse_ceru_excel(cell_map)
            if data is None:
                return {
                    "parser_type": PARSER_CERU,
                    "data": None,
                    "error": "Failed to parse CERU format: validation failed",
                }
            return {"parser_type": PARSER_CERU, "data": data}

        elif parser_type == PARSER_ZO_PAINT:
            data = parse_rfq_zo_paint_to_dict(
                file_path=file_path,
                sheet_name=sheet_name or "store req. form",
                parse_descriptions=True,
            )
            return {"parser_type": PARSER_ZO_PAINT, "data": data}

        elif parser_type == PARSER_INDONESIA:
            data = parse_rfq_excel_to_dict(
                file_path=file_path,
                header_row_index=None,
                required_column="Delivery time (in working days)",
            )
            return {"parser_type": PARSER_INDONESIA, "data": data}

        else:
            return {
                "parser_type": PARSER_UNKNOWN,
                "data": None,
                "error": f"Unknown parser type: {parser_type}",
            }

    except Exception as e:
        logger.error(f"Error parsing with {parser_type} parser: {e}", exc_info=True)
        return {"parser_type": parser_type, "data": None, "error": str(e)}


def get_supported_formats() -> list[dict[str, str]]:
    """Get list of supported Excel formats.

    Returns:
        List of dictionaries with format information.
    """
    return [
        {
            "type": PARSER_CERU,
            "name": "CERU Format",
            "description": "Cerulean vessel RFQ format with requisition details",
        },
        {
            "type": PARSER_ZO_PAINT,
            "name": "ZO Paint Format",
            "description": "Store request form with paint items and sections",
        },
        {
            "type": PARSER_INDONESIA,
            "name": "Indonesian RFQ Format",
            "description": "Indonesian vessel RFQ with Part Number and Description",
        },
    ]
