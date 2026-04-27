"""
Competitor Matrix Excel Reader

Utility module for reading and parsing competitor matrix Excel files.
Extracts generic products, brands, and their equivalent products.
"""

import logging
from dataclasses import dataclass
from typing import Any

import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)


@dataclass
class ExcelSheetStructure:
    """Represents the structure of an Excel sheet"""

    sheet_name: str
    cells: dict[str, Any]  # cell_address -> cell_value


@dataclass
class CompetitorMatrixData:
    """Parsed competitor matrix data"""

    generics: list[str]
    brands: list[str]
    products: list[dict[str, Any]]  # List of {generic, brand, product}
    default_df: pd.DataFrame  # Store the default DataFrame


class CompetitorMatrixReader:
    """
    Reader for competitor matrix Excel files.

    Expected Excel format:
    - First column: GENERIC (generic product names)
    - Other columns: Brand names with their equivalent products
    - Header row: Specified row number (e.g., row 5)
    """

    DEFAULT_HEADER_ROW = 5
    DEFAULT_GENERIC_COLUMN = "GENERIC"

    def __init__(self, header_row: int = None):
        """
        Initialize the reader.

        Args:
            header_row: The 1-based row number to use as headers (default: 5)
        """
        self.header_row = header_row or self.DEFAULT_HEADER_ROW

    def read_excel_file(self, file_path: str) -> dict[str, ExcelSheetStructure]:
        """
        Read an Excel file and return structure for all sheets.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary mapping sheet names to their structures

        Raises:
            FileNotFoundError: If file doesn't exist
            Exception: For other read errors
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except FileNotFoundError:
            raise FileNotFoundError(f"File not found at '{file_path}'")
        except Exception as e:
            raise Exception(f"Error loading workbook: {e}")

        all_sheet_structures = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_structure = {}
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        sheet_structure[cell.coordinate] = cell.value
            all_sheet_structures[sheet_name] = ExcelSheetStructure(
                sheet_name=sheet_name, cells=sheet_structure
            )

        return all_sheet_structures

    def parse_sheet_to_dataframe(
        self, sheet_structure: ExcelSheetStructure
    ) -> pd.DataFrame:
        """
        Convert a sheet structure to pandas DataFrame.

        Args:
            sheet_structure: The sheet structure to parse

        Returns:
            DataFrame with parsed data
        """
        sheet_data = sheet_structure.cells

        # Organize cells by row number
        organized_by_row = {}
        for cell_address, cell_value in sheet_data.items():
            col = "".join(filter(str.isalpha, cell_address))
            row = int("".join(filter(str.isdigit, cell_address)))
            if row not in organized_by_row:
                organized_by_row[row] = {}
            organized_by_row[row][col] = cell_value

        # Get header columns from the specified row
        header_cells = organized_by_row.get(self.header_row, {})
        sorted_cols = sorted(header_cells.keys())
        column_names = [header_cells[col] for col in sorted_cols]

        # Handle None column names
        final_column_names = [
            col_name if col_name is not None else f"Unnamed_Col_{i}"
            for i, col_name in enumerate(column_names)
        ]

        # Extract data rows after the header
        rows_list = []
        for r in sorted(organized_by_row.keys()):
            if r > self.header_row:
                row_values = []
                for col in sorted_cols:
                    row_values.append(organized_by_row[r].get(col))
                rows_list.append(row_values)

        # Create DataFrame
        df = pd.DataFrame(rows_list, columns=final_column_names)
        return df

    def parse_competitor_matrix(
        self, file_path: str, sheet_name: str = None
    ) -> CompetitorMatrixData:
        """
        Parse competitor matrix Excel file.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of sheet to parse (default: first sheet)

        Returns:
            CompetitorMatrixData with parsed generics, brands, and products
        """
        # Read Excel file
        all_sheets = self.read_excel_file(file_path)

        # Use first sheet if not specified
        if sheet_name is None:
            sheet_name = list(all_sheets.keys())[0]

        if sheet_name not in all_sheets:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")

        sheet_structure = all_sheets[sheet_name]

        # Convert to DataFrame
        df = self.parse_sheet_to_dataframe(sheet_structure)

        logger.info(f"Parsing sheet '{sheet_name}' with {len(df)} rows")

        logger.info(f"Available columns: {list(df.columns)}")

        logger.info(f"Header row: {self.header_row}")

        logger.info(f"Data sample:\n{df.head()}")

        # Validate required columns
        if self.DEFAULT_GENERIC_COLUMN not in df.columns:
            raise ValueError(
                f"Column '{self.DEFAULT_GENERIC_COLUMN}' not found in sheet. Available columns: {list(df.columns)}"
            )

        # Extract generics
        generics = (
            df[self.DEFAULT_GENERIC_COLUMN]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
            .tolist()
        )

        # Extract brands (all columns except GENERIC)
        brands = [col for col in df.columns if col != self.DEFAULT_GENERIC_COLUMN]

        # Extract products (mapping of generic -> brand -> product)
        products = []
        for _, row in df.iterrows():
            generic = row[self.DEFAULT_GENERIC_COLUMN]
            if pd.isna(generic) or str(generic).strip() == "":
                continue

            generic = str(generic).strip()

            for brand in brands:
                product = row.get(brand)
                if pd.notna(product) and str(product).strip() != "":
                    products.append(
                        {
                            "generic": generic,
                            "brand": brand,
                            "product": str(product).strip(),
                        }
                    )

        return CompetitorMatrixData(
            generics=generics, brands=brands, products=products, default_df=df
        )

    def get_available_sheets(self, file_path: str) -> list[str]:
        """
        Get list of available sheet names in Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            List of sheet names
        """
        all_sheets = self.read_excel_file(file_path)
        return list(all_sheets.keys())


def parse_competitor_matrix_excel(
    file_path: str, sheet_name: str = None, header_row: int = None
) -> CompetitorMatrixData:
    """
    Convenience function to parse competitor matrix Excel file.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of sheet to parse (default: first sheet)
        header_row: The 1-based row number to use as headers (default: 5)

    Returns:
        CompetitorMatrixData with parsed data
    """
    reader = CompetitorMatrixReader(header_row=header_row)
    return reader.parse_competitor_matrix(file_path, sheet_name)
